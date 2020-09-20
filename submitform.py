import os
import xlwt
import mechanize
import xlsxwriter 
import openpyxl as op
import pandas as pd
from xlwt import Workbook
from bs4 import BeautifulSoup

def write_contexts(data):
    print('\n\t\t --------- Writing to EXCEL ------------')
    wb = op.load_workbook('AirUniversityApplicantData', False)
    ws = wb['Sheet1']
    ws.cell(row=ws.max_row, column=7).value += data
    wb.save('AirUniversityApplicantData')
    wb.close()

def read_excel():
    df = pd.read_excel('AirUniversityApplicantData.xlsx')
    card = df['Admit Card']
    name = df['Full Name']
    father = df['Father Name']
    return card, name, father

def select_form(form):
  return form.attrs.get('id', None) == 'form1'

def write_text(realdiv, card, name, father):
    with open('data.txt', 'a') as f:
        f.write('Admit Card   : ' + str(card) + '\n')
        f.write('Student Name : ' + str(name) + '\n')
        f.write('Father Name  : ' + str(father) + '\n')
        f.write(realdiv.text)
        f.write('----------------------------------------------------------------------------\n')

def get_data(card):
    br = mechanize.Browser()
    br.open('http://portals.au.edu.pk/result/')

    br.select_form(predicate=select_form)
    br.form["ctl00$AUContent$txt_regid"] = str(card)
    br.form.set_all_readonly(False)
    br.submit()

    return br.response().read()

if __name__ == "__main__":
        
    card, name, father = read_excel()
    
    for index in range(len(card)):
        print(card[index], '\t', name[index], '\t', father[index])
        source = get_data(card[index])
        soup = BeautifulSoup(source, features='lxml')
        mydiv = soup.find_all('div', attrs = {'class' : 'card h-100'})

        realdiv = None
        for div in mydiv:   
            realdiv = div
        write_text(realdiv, card[index], name[index], father[index] )

